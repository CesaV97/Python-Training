import openpyxl

#Codigo para crear un fichero
wb = openpyxl.Workbook()
wb.create_sheet("hoja1")
wb.save("Prueba.xlsx")
wb.close()

wb1 = openpyxl.load_workbook("C:\\Users\\Cesar Vega\\Desktop\\Phyton\\Prueba.xlsx")
hoja = wb1.active
hoja["A1"] = "Valor"
hoja.cell(row=2,column=2).value = 10
print(hoja["A1"].value)
print(hoja.cell(row=2,column=2).value)

print(hoja.max_column)
print(hoja.max_row)

wb1.save("Prueba.xlsx")
wb1.close()