"""
*En un centro educativo, los datos y calificaciones del
alumnado se guarda en un fichero de nombre
calificaciones.xlsx. la primera columna contiene los
apeelidos y nombres de los alumnos en orden alfabetico,
las dos siguientes columnas las notas de los dos parciales
teoricos realizados y la ultima columna las notas de practicas

*sabiendo que el peso de cada parcial teorico en la nota final
es de 30% y las notas de practicas cuentan un 40%, desarrolla
una funcion que a partir del fichero de calificaciones, devuelva
un diccionario donde las claves son los apellidos y nombre del
alumnado y los valores la nota final calculada
"""

import openpyxl #importamos la libreria para excel

# creamos la funcion
def calcular_notas(fichero):
    excel = openpyxl.load_workbook(fichero) #abrimos el archivo 
    hoja = excel.active # vemos la hoja activa
    filas = hoja.max_row # calculamos el valor de las filas de la hojanexcel
    notas = {} # creamos la variable notas de tipo Diccionario
    
    #fila toma el valor de iteraciones dependiendo a la cantidad de filas en el rango 1 a cantidad de filas
    for fila in range(1,filas): #Va de 1 a 6 + 1 = 7
        nombre = hoja.cell(row=fila, column=1).value # extraemos los calores de cada fila y columna
        nota_teorica01 = hoja.cell(row=fila, column=2).value
        nota_teorica02 = hoja.cell(row=fila, column=3).value
        nota_practica = hoja.cell(row=fila, column=4).value    
        
        nota_final = nota_teorica01 *0.3 + nota_teorica02 *0.3 + nota_practica *0.4 # calculamos el valor
        notas[nombre] = round(nota_final,2) #creamos la variable notas en en donde toma el valor
        #de la variable nombre y el valor de la nota final redondeando el valor con 2 decimales
        
    return notas # regresamos el valor de la variable notas de tipo diccionario 
    
#Se imprimi el valor en diccionario [nombre]:nota_final    
print(calcular_notas("C:\\Users\\Cesar Vega\\Desktop\\Phyton\\calificaciones.xlsx"))